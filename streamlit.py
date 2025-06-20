# streamlit_app.py

import streamlit as st
import tempfile, os
from pathlib import Path

# ——————————————————————————————
# 1) Import all the libraries you need
#    (same as your script: pandas, numpy, scipy, xgboost, openpyxl, etc.)
# ——————————————————————————————
import pandas as pd
import numpy as np
import openpyxl
import xgboost as xgb
#import pmdarima as pm
from scipy.stats import norm, gamma, kstest
from sklearn.model_selection import TimeSeriesSplit, GridSearchCV
from sklearn.metrics import mean_squared_error, mean_absolute_error
import matplotlib.pyplot as plt
from openpyxl.utils import column_index_from_string, get_column_letter

# …any other imports you used (e.g. seaborn if you still want plots in the notebook)…
# Note: Streamlit can show matplotlib charts directly, but if you only need Excel output,
#       you don’t have to display the charts in the Streamlit UI.


# ——————————————————————————————
# 2) Wrap your existing script into one function
#    that reads an input Excel and writes a new Excel with forecasts.
#    You can literally paste your code inside here, replacing hardcoded filenames
#    with the in_path/out_path parameters.
# ——————————————————————————————

def run_forecast_pipeline(in_path: str, out_path: str):
    """
    1) Load 'GLOBAL_LEAN_1_copia.xlsx' from in_path
    2) Do all the sheet-by-sheet loops, compute ARIMA, XGBoost, Monte Carlo, etc.
    3) Write the final Forecast sheet and any Heijunka formulas into out_path
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(in_path)
    wb = openpyxl.load_workbook(filename=in_path)
    sheet_ins = workbook["Instrucciones"]
    valor = sheet_ins["C2"].value

    # Create dataframes for each "Proceso i"
    dfs = {}
    params_dict = {}
    for i in range(1, valor + 1):
        df_i = pd.read_excel(in_path, sheet_name=f"Proceso {i}")
        sheet_i = workbook[f"Proceso {i}"]

        turno   = sheet_i["H1"].value
        horas   = sheet_i["H2"].value
        descanso= sheet_i["H3"].value
        prod    = sheet_i["H4"].value
        prod_obj= sheet_i["H6"].value
        trab    = sheet_i["H7"].value

        # Keep only the needed columns:
        df_i = df_i[["Día", "Turno", "Carga histórica", "Número de FTE"]].copy()
        # Add Date column, drop original Día
        df_i["Date"] = pd.to_datetime(df_i["Día"])
        df_i.drop(columns=["Día"], inplace=True)
        # Group stats (mean/std) if you want plots or stats
        # ...
        # Save to dict for later use:
        #dfs[i] = df_i.copy()
        #params_dict[i] = dict(
        #    turno=turno,
        #    horas=horas,
        ##    descanso=descanso,
        #    productividad=prod,
        #    productividad_objetivo=prod_obj,
        #    trabajadores=trab
        #)

    # Example: Focus on Proceso 1 for demonstration (you can loop all i)
    #df1 = dfs[1]
    #params1 = params_dict[1]

    # Split train/test by last month
        last_month = df_i['Date'].dt.to_period("M").max()
        df1_train = df_i[df_i['Date'].dt.to_period("M") < last_month].copy()
        df1_test  = df_i[df_i['Date'].dt.to_period("M") == last_month].copy()

    # ——————————————————————————————
    # 3) MONTE CARLO SIMULATION (if you want that)
    #    (you already had a function `simulacion_montecarlo(df)` in your script)
    # ——————————————————————————————
        def create_features(df, label=None):
            df2 = df.copy()
            df2["dayofweek"] = df2["Date"].dt.dayofweek
            df2["quarter"]   = df2["Date"].dt.quarter
            df2["month"]     = df2["Date"].dt.month
            df2["year"]      = df2["Date"].dt.year
            df2["dayofyear"] = df2["Date"].dt.dayofyear
            df2["day"]       = df2["Date"].dt.day
            df2["week"]      = df2["Date"].dt.isocalendar().week

            X = df2[["dayofweek", "quarter", "month", "year", "dayofyear", "day", "week"]]
            if label:
                y = df2[label]
                return X, y
            return X

        # Train XGBoost on df1_train
        X_train, y_train = create_features(df1_train, label="Carga histórica")
        X_test,  y_test  = create_features(df1_test,  label="Carga histórica")
    
    #xgb_model = xgb.XGBRegressor(objective='reg:squarederror', random_state=42)
    #param_grid = {
    #      'n_estimators': [50, 200, 500],
    #      'max_depth':    [1, 5, 10],
    #      'learning_rate':[0.005, 0.05, 0.15]
    #  }

    #print("Hello")

  # 5. TimeSeriesSplit + GridSearchCV
    #tscv = TimeSeriesSplit(n_splits=5)
    #print("Hello1")
    #model = GridSearchCV(
    #    estimator=xgb_model,
    #    param_grid=param_grid,
    #    cv=tscv,
    #    scoring='neg_mean_squared_error',
    #    verbose=False,
    #    n_jobs=-1
    #  )
    #print("Hello2")
    #model.fit(X_train, y_train)
    #print("Hello3")
        xgb_model = xgb.XGBRegressor(n_estimators=200, max_depth=1, learning_rate=0.005, subsample=0.8)
        xgb_model.fit(X_train, y_train)
        # Forecast next 30 days
        last_date = df_i["Date"].max()
        future_dates = pd.date_range(start=last_date + pd.Timedelta(days=1), periods=30, freq="D")
        df_future = pd.DataFrame({"Date": future_dates})
        X_future = create_features(df_future)
        preds_future = xgb_model.predict(X_future)

    # ——————————————————————————————
    # 4) WRITE THE NEW "Forecast" SHEET into the same workbook object
    # ——————————————————————————————
        if f"Forecast {i}" in workbook.sheetnames:
            del workbook[f"Forecast {i}"]
        sheet_fc = workbook.create_sheet(title=f"Forecast {i}")
        bold_font = openpyxl.styles.Font(bold=True)

    # Build a DataFrame for these predictions:
        df_final = pd.DataFrame({
            "Día": future_dates,
            "Demanda": preds_future
        })

        # Write headers with bold:
        sheet_fc["A1"].value = "Día"
        sheet_fc["A1"].font  = bold_font
        sheet_fc["B1"].value = "Demanda"
        sheet_fc["B1"].font  = bold_font
        sheet_fc["C1"].value = "Edición"
        sheet_fc["C1"].font  = bold_font
        sheet_fc["D1"].value = "Demanda Final"
        sheet_fc["D1"].font  = bold_font

        # Write data rows:
        for row_idx, row in enumerate(df_final.itertuples(index=False), start=2):
            sheet_fc.cell(row=row_idx, column=1).value = row.Día
            sheet_fc.cell(row=row_idx, column=2).value = float(row.Demanda)
            # Column C and D are left for user “Edición” / “Demanda Final” formulas
            # You can also pre-fill them with formulas if you want:
            sheet_fc.cell(row=row_idx, column=4).value = f"=IF(C{row_idx}=0, B{row_idx}, C{row_idx})"

        # Fill editable cells in column C with a yellow background:
        fill_yellow = openpyxl.styles.PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid")
        for r in range(2, 2 + len(df_final)):
            sheet_fc.cell(row=r, column=3).fill = fill_yellow

    # Add FTE columns E/F/G if you need those formulas for “Proceso 1”
        trabajadores = trab
    #prod        = params1["productividad"]
    #prod_obj    = params1["productividad_objetivo"]

        sheet_fc["E1"].value = "Número de FTE actuales";      sheet_fc["E1"].font = bold_font
        sheet_fc["F1"].value = "Productividad";               sheet_fc["F1"].font = bold_font
        sheet_fc["G1"].value = "Productividad objetivo";      sheet_fc["G1"].font = bold_font
        sheet_fc["I1"].value = "Heijunka 1"
        sheet_fc["I1"].font = bold_font
        sheet_fc["J1"].value = "Horas necesarias"
        sheet_fc["J1"].font = bold_font
        sheet_fc["K1"].value = "FTE necesarias"
        sheet_fc["K1"].font = bold_font
        sheet_fc["L1"].value = "Horas Objetivo"
        sheet_fc["L1"].font = bold_font
        sheet_fc["M1"].value = "FTE objetivos"
        sheet_fc["M1"].font = bold_font
        sheet_fc["N1"].value = "Diferencia Horas"
        sheet_fc["N1"].font = bold_font
        sheet_fc["O1"].value = "Diferencia FTE"
        sheet_fc["O1"].font = bold_font
        sheet_fc["P1"].value = "Ocupación (Porcentaje)"
        sheet_fc["P1"].font = bold_font
        sheet_fc["Q1"].value = "Exceso/Falta de Horas (horas)"
        sheet_fc["Q1"].font = bold_font
        sheet_fc["R1"].value = "Coste hora extra (euros)"
        sheet_fc["R1"].font = bold_font
        sheet_fc["S1"].value = "Coste hora ociosa (euros)"
        sheet_fc["S1"].font = bold_font
        sheet_fc["T1"].value = "Coste ineficiente (euros)"
        sheet_fc["T1"].font = bold_font

        for r in range(2, 2 + len(df_final)):
            sheet_fc.cell(row=r, column=4).value = f"=IF(C{r} = 0, B{r}, C{r})"
            sheet_fc.cell(row=r, column=5).value = f"='Proceso 1'!$H$7"
            sheet_fc.cell(row=r, column=6).value = f"='Proceso 1'!$H$4"
            sheet_fc.cell(row=r, column=7).value = f"='Proceso 1'!$H$6"
            sheet_fc.cell(row=r, column=10).value = f"=D{r} / F{r}"
            sheet_fc.cell(row=r, column=11).value = f"=J{r}/('Proceso 1'!$H$2-('Proceso 1'!$H$3/60))"
            sheet_fc.cell(row=r, column=11).font = bold_font
            sheet_fc.cell(row=r, column=12).value = f"=D{r} / G{r}"
            sheet_fc.cell(row=r, column=13).value = f"=L{r}/('Proceso 1'!$H$2-('Proceso 1'!$H$3/60))"
            sheet_fc.cell(row=r, column=14).value = f"=J{r} - L{r}"
            sheet_fc.cell(row=r, column=15).value = f"=K{r} - M{r}"
            sheet_fc.cell(row=r, column=16).value =  f"=J{r}/(E{r}*('Proceso 1'!$H$2-('Proceso 1'!$H$3/60)))"
            sheet_fc.cell(row=r, column=16).font = bold_font
            sheet_fc.cell(row=r, column=17).value = f"=J{r} - (E{r}*('Proceso 1'!$H$2-('Proceso 1'!$H$3/60)))"
            sheet_fc.cell(row=r, column=18).value = 25
            sheet_fc.cell(row=r, column=19).value = 18
            sheet_fc.cell(row=r, column=20).value = f"=IF(Q{r}>0,Q{r}*R{r},-1*Q{r}*S{r})"
            sheet_fc.cell(row=r, column=20).font = bold_font

        sheet_fc["V1"].value = "Heijunka 2"
        sheet_fc["V1"].font = bold_font
        sheet_fc["W1"].value = "Número de FTE propuesto"
        sheet_fc["W1"].font = bold_font
        sheet_fc["X1"].value = "Exceso/Falta de Horas (horas)"
        sheet_fc["X1"].font = bold_font
        sheet_fc["Y1"].value = "Coste hora extra (euros)"
        sheet_fc["Y1"].font = bold_font
        sheet_fc["Z1"].value = "Coste hora ociosa (euros)"
        sheet_fc["Z1"].font = bold_font
        sheet_fc["AA1"].value = "Coste ineficiente (euros)"
        sheet_fc["AA1"].font = bold_font
        sheet_fc["AB1"].value = "Coste total (euros)"
        sheet_fc["AB1"].font = bold_font

        sheet_fc["AD3"].value = 1

        for row in range(2, 2 + len(df_final)):
            sheet_fc.cell(row = row, column = 23).value = f"=$AD$3"
            sheet_fc.cell(row=row, column=24).value = f"=J{row} - (W{row}*('Proceso 1'!$H$2-('Proceso 1'!$H$3/60)))"
            sheet_fc.cell(row=row, column=25).value = 25
            sheet_fc.cell(row=row, column=26).value = 18
            sheet_fc.cell(row=row, column=27).value = f"=IF(X{row}>0,X{row}*Y{row},-1*X{row}*Z{row})"

        sheet_fc["AD4"].value = "=SUM(AA2:AA31)"
        sheet_fc["AD4"].font = bold_font

        ws = workbook[f"Proceso {i}"]

        H2= ws["H2"].value    # e.g. 8
        H3= ws["H3"].value
        R, S   = 25.0, 18.0
        avail = H2 - (H3 / 60.0)
        J = df_final["Demanda"].values / prod

        def total_cost(W: float) -> float:
            X  = J - W * avail
            # piecewise cost
            AA = np.where(X > 0, X * R, -X * S)
            return AA.sum()

        Ws = np.arange(0, 5001, 1)      # try W=0.0,0.1,0.2,...,50.0
        costs = np.array([total_cost(w) for w in Ws])
        idx   = costs.argmin()
        
        W_opt    = Ws[idx]
        min_cost = costs[idx]
        
        print(f"Optimal staff (W) = {W_opt:.2f}")
        print(f"Minimum total cost = {min_cost:.2f}")
    
        sheet_fc["AD3"].value  = W_opt
        sheet_fc["AD3"].font = bold_font
        sheet_fc["AC3"].value = "Número de FTE en el mes"
        sheet_fc["AC3"].font = bold_font
        sheet_fc["AC4"].value = "Sobrecoste en el mes (euros)"
        sheet_fc["AC4"].font = bold_font

        def _is_formula(cell):
            """
            Returns True if the cell holds an Excel formula string (i.e. its value is a str starting with '=').
            """
            return isinstance(cell.value, str) and cell.value.startswith("=")
        
        # 2) Loop over all rows, skip the header row (row 1)
        for row in sheet_fc.iter_rows(min_row=2, max_row=sheet_fc.max_row, 
                                    min_col=2, max_col=sheet_fc.max_column):
            for cell in row:
                val = cell.value
        
                # (a) If it’s blank or None, do nothing
                if val is None:
                    continue
        
                # (b) If it’s a formula (starts with "="), just apply two‐decimal number_format
                if _is_formula(cell):
                    cell.number_format = "0.00"
                    continue
        
                # (c) Otherwise, try to coerce into float (e.g. "123.456" or numeric)
                try:
                    num = float(val)
                except Exception:
                    # couldn’t parse as float → leave as-is (e.g. a text field, or something non-numeric)
                    continue
                else:
                    # we have a numeric value; round it and re-assign
                    cell.value = round(num, 2)
                    cell.number_format = "0.00"
                    
        last_row = 1 + len(df_final)

        pink_fill   = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_fill   = openpyxl.styles.PatternFill(start_color="5379F3", end_color="5379F3", fill_type="solid")

        for r in range(2, last_row + 1):
            # Columns I (col 9) through T (col 20) → pink
            for col_idx in range(openpyxl.utils.column_index_from_string("I"), openpyxl.utils.column_index_from_string("T") + 1):
                sheet_fc.cell(row=r, column=col_idx).fill = pink_fill
        
            # Columns V (col 22) through AD (col 30) → blue
            for col_idx in range(openpyxl.utils.column_index_from_string("V"), openpyxl.utils.column_index_from_string("AD") + 1):
                sheet_fc.cell(row=r, column=col_idx).fill = blue_fill

    # ——————————————————————————————
    # 7) Auto‐size columns and rows in “Forecast” sheet
    # ——————————————————————————————
    # (A) Auto‐size each column by its max text length
    # 3) Auto-size columns A…Z… (but skip raw formula lengths and clamp to MAX_COL_WIDTH)
        for col_cells in sheet_fc.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            # Add a little extra padding (e.g. +2 characters)
            sheet_fc.column_dimensions[col_letter].width = max_length


        # (B) Auto‐size each row by number of wrapped lines
        for row_cells in sheet_fc.rows:
            max_lines = 1
            for cell in row_cells:
                if cell.value is not None and isinstance(cell.value, str):
                    lines = str(cell.value).split("\n")
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            sheet_fc.row_dimensions[row_cells[0].row].height = max_lines * 15


        # … you can keep adding all of your Heijunka grid/formula logic here, 
        #    e.g. columns H..Z with formulas referencing cells in "Proceso 1" …
        #    (just translate each f"=…" Excel formula into sheet_fc.cell(row=…, col=…).value = "…")

        # Finally, save the updated workbook to out_path
        #workbook.save(out_path)
        
    sheet_final = workbook.create_sheet(title="Heijunka 3")
    bold_font = openpyxl.styles.Font(bold=True)
    sheet_final["A1"].value = "Heijunka 3"
    sheet_final["A1"].font = bold_font
    sheet_final["B2"].value = "Proceso"
    sheet_final["B2"].font = bold_font
    sheet_final["C2"].value = "Número de FTE actuales"
    sheet_final["C2"].font = bold_font
    sheet_final["D2"].value = "Número de FTE propuesto"
    sheet_final["D2"].font = bold_font
    sheet_final["E2"].value = "Diferencia FTE"
    sheet_final["E2"].font = bold_font

    for i in range(1, valor + 1):
        num = 0
        sheet_final.cell(row=i + 2, column=2).value = f"Proceso {i}"
        sheet_final.cell(row=i + 2, column=3).value = f"='Proceso {i}'!$H$7"
        sheet_final.cell(row=i + 2, column=4).value = f"='Forecast {i}'!$AD$3"
        sheet_final.cell(row=i + 2, column=5).value = f"='Forecast {i}'!$AD$3-'Proceso {i}'!$H$7"
        if i > num: 
            num = i

    sheet_final.cell(row=num + 3, column=5).value = f"=SUM(E3:E{num + 2})"
    sheet_final.cell(row=num + 3, column=5).font = bold_font
    sheet_final.cell(row=num + 3, column=5).fill = pink_fill

    workbook.save(out_path)

# ——————————————————————————————
# 3) Streamlit UI: file uploader + call pipeline + download butt
# ——————————————————————————————
st.title("Herramienta Heijunka Global Lean")
st.markdown("""
            Bienvenido a la Herramienta Heijunka de Global Lean, donde podra obtener una previsión de la demanda para los proximos 30 días con datos históricos.
            Para ello, siga las siguientes instrucciones: \n
            1. Descargue la plantilla excel \n
            2. Rellene el excel con sus datos, siguiendo las instrucciones indicadas en la hoja de "Instrucciones" \n
            3. Introduzca el archivo en el recuadro de debajo y haga click en el boton para empezar la previsión \n
            4. Descargue el excel y ya tendrá su forecast. \n
            """)

RUTA_PLANTILLA = Path(__file__).parent / "assets" / "plantilla.xlsx"

# 2) Leo el archivo en modo binario
with open(RUTA_PLANTILLA, "rb") as f:
    plantilla_bytes = f.read()

st.download_button(
    label="📥 Descargar plantilla de Excel",
    data=plantilla_bytes,
    file_name="plantilla.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.markdown(
    """
    \n
    Introduzca aqui su excel. Se ejecutara el archivo Python 
    (XGBoost, ARIMA, Monte Carlo, etc.), y después podrá descargar su excel completo.
    """
)

uploaded_file = st.file_uploader("1) Adjunte su excel (.xlsx)", type=["xlsx"])
if uploaded_file:
    st.write(f"Elija el archivo: **{uploaded_file.name}**")

    # When the user clicks “Run Forecast,” save the uploaded contents to a temp file
    if st.button("2) Correr Forecast"):
        with st.spinner("Corriendo el archivo python…"):
            # Save upload to a temporary location
            tmp_input = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp_input.write(uploaded_file.read())
            tmp_input.flush()
            tmp_input_path = tmp_input.name

            # Decide where to write the output
            tmp_output_path = os.path.join(tempfile.gettempdir(), "Forecast_Demanda.xlsx")

            try:
                run_forecast_pipeline(tmp_input_path, tmp_output_path)
            except Exception as e:
                st.error(f"❌ Forecast pipeline failed:\n```\n{e}\n```")
            else:
                st.success("✅ Previsión completada! Descargue debajo:")
                # Read the output file into memory
                with open(tmp_output_path, "rb") as f:
                    data = f.read()
                # Provide a Streamlit download button
                st.download_button(
                    label="Descargue Heijunka Excel",
                    data=data,
                    file_name="Forecast_Demanda.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # Clean up the temp input file
            try:
                os.unlink(tmp_input_path)
            except OSError:
                pass
